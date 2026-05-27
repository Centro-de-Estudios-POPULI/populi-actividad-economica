"""
Genera charts HTML (ECharts) para indicadores adelantados de actividad economica.
Formato POPULI Design System con year-by-line (seasonal overlay) para mensuales.

Fuente: INE Bolivia (Excel)
Salida: embed/*.html  (6 archivos standalone)

Indicadores:
  9. Produccion de cemento           -> cemento_produccion.html
 10. Permisos de construccion        -> permisos_construccion.html
 11. Manufactura: produccion + cap.  -> manufactura_produccion_capacidad.html
 12. Consumo de energia electrica    -> energia_consumo.html
 13. Transporte indice general       -> transporte_indice.html
 14. Carga ferroviaria               -> transporte_carga_ferroviaria.html
"""

import json
import math
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

# -- Rutas ----------------------------------------------------------------
BASE = Path(__file__).resolve().parent.parent
DATOS = BASE / "datos" / "indicadores_adelantados"
EMBED = BASE / "embed"
EMBED.mkdir(parents=True, exist_ok=True)

# Meses en espanol -> numero
MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}
MESES_CORTO = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
               'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

TRIMESTRES = {"i": 1, "ii": 2, "iii": 3, "iv": 4}

# Year-by-line color palette (index 0 = most recent / current year)
YEAR_COLORS = ['#0D9488', '#2563EB', '#6366F1', '#D97706', '#94A3B8', '#CBD5E1', '#E2E8F0']
YEAR_WIDTHS = [3, 2.5, 2, 1.8, 1.5, 1.2, 1]
YEAR_COLORS_DARK = ['#0D9488', '#3B82F6', '#818CF8', '#F59E0B', '#94A3B8', '#64748B', '#475569']

# Special: 2020 dashed red
COVID_YEAR = 2020
COVID_COLOR = '#EF4444'
COVID_COLOR_DARK = '#F87171'


# -- Helpers de parseo ----------------------------------------------------

def _extract_year(val):
    """Extrae ano de un valor como 2025, 2025.0, '2025(p)', '2026 (p)'. Retorna int o None."""
    if val is None:
        return None
    try:
        y = int(float(val))
        if 1980 <= y <= 2040:
            return y
        return None
    except (ValueError, TypeError):
        pass
    s = str(val).strip()
    m = re.match(r'^(\d{4})\s*\(?p?\)?', s)
    if m:
        y = int(m.group(1))
        if 1980 <= y <= 2040:
            return y
    return None


def _month_num(val):
    """Devuelve numero de mes (1-12) o None."""
    if val is None:
        return None
    s = str(val).strip().lower()
    return MESES.get(s)


def _quarter_num(val):
    """Devuelve numero de trimestre (1-4) o None."""
    if val is None:
        return None
    s = str(val).strip().lower()
    for roman, num in TRIMESTRES.items():
        cleaned = s.lstrip()
        if cleaned.startswith(roman + " ") or cleaned.startswith(roman + "\t"):
            after = cleaned[len(roman):]
            if after.lstrip().startswith("trim"):
                return num
    for roman, num in sorted(TRIMESTRES.items(), key=lambda x: -len(x[0])):
        if roman + " " in s and "trim" in s:
            return num
    return None


def _safe_float(val):
    """Convierte a float o devuelve None."""
    if val is None:
        return None
    try:
        v = float(val)
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except (ValueError, TypeError):
        return None


def parse_monthly_xlsx(filepath, sheet_name, col_periodo, col_map, start_year=None):
    """
    Parsea archivo INE mensual (openpyxl).
    Retorna: dict {year: {month_idx(1-12): value}}  para cada serie en col_map.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    result = {name: {} for name in col_map}  # {serie: {year: {month: val}}}
    current_year = None

    for row in ws.iter_rows(min_row=1, values_only=False):
        period_cell = row[col_periodo - 1].value
        yr = _extract_year(period_cell)
        if yr is not None:
            current_year = yr
            continue
        if current_year is None:
            continue
        mn = _month_num(period_cell)
        if mn is None:
            continue
        if start_year and current_year < start_year:
            continue

        for name, col_idx in col_map.items():
            val = _safe_float(row[col_idx - 1].value)
            if current_year not in result[name]:
                result[name][current_year] = {}
            result[name][current_year][mn] = val

    wb.close()
    return result


def parse_monthly_xls(filepath, sheet_index, col_periodo, col_map, start_year=None):
    """
    Parsea archivo INE mensual .xls (via pandas/xlrd).
    Retorna: dict {serie: {year: {month: val}}}
    """
    df = pd.read_excel(str(filepath), sheet_name=sheet_index, engine="xlrd", header=None)

    result = {name: {} for name in col_map}
    current_year = None

    for _, row in df.iterrows():
        period_val = row.iloc[col_periodo]
        yr = _extract_year(period_val)
        if yr is not None:
            current_year = yr
            continue
        if current_year is None:
            continue
        mn = _month_num(period_val)
        if mn is None:
            continue
        if start_year and current_year < start_year:
            continue

        for name, col_idx in col_map.items():
            val = _safe_float(row.iloc[col_idx])
            if current_year not in result[name]:
                result[name][current_year] = {}
            result[name][current_year][mn] = val

    return result


def parse_quarterly_xlsx(filepath, sheet_name, col_periodo, col_map, start_year=None):
    """Parsea archivo INE trimestral (openpyxl). Retorna (labels, {serie: [values]})."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    labels = []
    series = {name: [] for name in col_map}
    current_year = None

    for row in ws.iter_rows(min_row=1, values_only=False):
        period_cell = row[col_periodo - 1].value
        yr = _extract_year(period_cell)
        if yr is not None:
            current_year = yr
            continue
        if current_year is None:
            continue
        qn = _quarter_num(period_cell)
        if qn is None:
            continue
        if start_year and current_year < start_year:
            continue

        labels.append(f"{current_year}-T{qn}")
        for name, col_idx in col_map.items():
            val = _safe_float(row[col_idx - 1].value)
            series[name].append(val)

    wb.close()
    return labels, series


def json_safe(values):
    """Convierte lista de floats/None a lista JSON-safe."""
    return [round(v, 2) if v is not None else None for v in values]


def build_year_data(year_dict, years):
    """
    Dado {year: {month: val}} y lista de anos,
    retorna lista de (year, [val_ene, val_feb, ..., val_dic])
    donde val puede ser None.
    """
    rows = []
    for y in years:
        vals = []
        if y in year_dict:
            for m in range(1, 13):
                vals.append(year_dict[y].get(m))
        else:
            vals = [None] * 12
        rows.append((y, vals))
    return rows


# -- HTML Template (POPULI Design System) ---------------------------------

def _html_template(title, subtitle, chart_js, accent_color="#0D9488"):
    """Genera HTML completo con POPULI design system."""
    return f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>{title}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300..800&family=Playfair+Display:ital,wght@0,400..900;1,400..900&family=JetBrains+Mono:wght@400;500;600;700&display=swap" rel="stylesheet"/>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
:root{{
  --populi:#8B1A1A;--populi-dark:#6B0300;--populi-light:#C00000;--populi-gold:#D4A017;
  --cream:#F5EFE0;--brown:#5C3D1E;--brown-dark:#3D2B1F;
  --navy:#0D1B2A;--navy-light:#1A2940;
  --slate:#475569;--slate-light:#64748B;
  --warm-white:#FAF8F3;--light-gray:#F1EDE5;
  --border:#E2DDD3;--dark-border:#2A3A50;
  --bg:var(--warm-white);--card:#FFFFFF;--text:var(--brown-dark);--muted:var(--slate-light);
  --radius:12px;
}}
[data-theme="dark"]{{
  --bg:#080808;--card:#141414;--text:#E2E8F0;--muted:#64748B;
  --border:#2A3A50;--light-gray:#1A1A1A;--cream:#1A2940;
}}
body{{
  font-family:'Inter',system-ui,sans-serif;
  background:var(--bg);color:var(--text);
  -webkit-font-smoothing:antialiased;
  transition:background .25s,color .25s;
}}
.embed-container{{max-width:1200px;margin:0 auto;padding:20px 16px}}
.chart-card{{
  background:var(--card);border:1px solid var(--border);border-radius:var(--radius);
  padding:16px;display:flex;flex-direction:column;overflow:hidden;
}}
[data-theme="dark"] .chart-card{{background:#141414;border-color:var(--dark-border)}}
.section-title{{
  font-family:'Playfair Display',Georgia,serif;font-size:1.55rem;font-weight:700;
  color:var(--navy);display:flex;align-items:center;gap:10px;
}}
[data-theme="dark"] .section-title{{color:#fff}}
.accent-bar{{width:4px;height:26px;border-radius:2px;flex-shrink:0;background:{accent_color}}}
.section-subtitle{{font-size:.78rem;color:var(--muted);padding-left:14px;margin-top:2px}}
#chart{{width:100%;height:480px;margin-top:12px}}
.chart-source{{
  display:flex;align-items:center;justify-content:space-between;
  padding:10px 4px 0;margin-top:8px;border-top:1px solid var(--border);
}}
.chart-source-text{{font-size:.72rem;color:var(--muted);letter-spacing:.01em}}
.chart-source-text a{{color:var(--populi);text-decoration:none;font-weight:600}}
[data-theme="dark"] .chart-source-text a{{color:#E9D8A6}}
.watermark{{width:80px;height:auto;opacity:.18}}
[data-theme="dark"] .watermark text:first-child{{fill:#E9D8A6}}
[data-theme="dark"] .watermark text:last-child{{fill:#94A3B8}}
</style>
</head>
<body>
<div class="embed-container">
  <div class="chart-card">
    <div class="section-title">
      <span class="accent-bar"></span>{title}
    </div>
    <div class="section-subtitle">{subtitle}</div>
    <div id="chart"></div>
    <div class="chart-source">
      <span class="chart-source-text">Fuente: <a>INE</a> &middot; Elaboraci&oacute;n: <a>Centro de Estudios POPULI</a></span>
      <svg viewBox="0 0 310 130" class="watermark"><text x="6" y="112" font-family="'Playfair Display'" font-size="145" font-weight="700" fill="#8B1A1A">P</text><text x="92" y="87" font-family="'Playfair Display'" font-size="55" font-weight="400" font-style="italic" fill="#3D2B1F">opuli</text></svg>
    </div>
  </div>
</div>
<script>
(function(){{
  var params=new URLSearchParams(window.location.search);
  if(params.get('theme')==='dark') document.documentElement.dataset.theme='dark';
  window.addEventListener('message',function(e){{if(e.data&&e.data.theme) document.documentElement.dataset.theme=e.data.theme}});
  function dk(){{return document.documentElement.dataset.theme==='dark'}}

  var chartDom=document.getElementById('chart');
  var chart=echarts.init(chartDom);

  function buildChart(){{
{chart_js}
    chart.setOption(option,true);
  }}

  buildChart();
  new MutationObserver(function(){{buildChart()}}).observe(document.documentElement,{{attributes:true,attributeFilter:['data-theme']}});
  window.addEventListener('resize',function(){{chart.resize()}});
}})();
</script>
</body>
</html>'''


# -- Year-by-line series builder (JS) ------------------------------------

def build_year_line_series_js(year_data_list, unit="", format_fn="null"):
    """
    Genera el JS para las series de year-by-line.
    year_data_list: [(year, [12 vals]), ...] ordered newest-first
    """
    series_js_parts = []
    for i, (year, vals) in enumerate(year_data_list):
        is_covid = (year == COVID_YEAR)
        color_idx = min(i, len(YEAR_COLORS) - 1)
        width_idx = min(i, len(YEAR_WIDTHS) - 1)

        if is_covid:
            color_light = COVID_COLOR
            color_dark = COVID_COLOR_DARK
            line_type = "'dashed'"
            width = 1.8
        else:
            color_light = YEAR_COLORS[color_idx]
            color_dark = YEAR_COLORS_DARK[color_idx]
            width = YEAR_WIDTHS[width_idx]
            line_type = "'solid'"

        opacity = max(0.3, 1.0 - i * 0.1) if not is_covid else 0.7
        vals_json = json.dumps(json_safe(vals))

        # Find last non-null index for markPoint on current year (i==0)
        mark_point_js = ""
        if i == 0:
            last_idx = -1
            for mi in range(11, -1, -1):
                if vals[mi] is not None:
                    last_idx = mi
                    break
            if last_idx >= 0 and last_idx < 11:
                mark_point_js = f"""
        markPoint:{{
          data:[{{
            coord:[{last_idx},{json.dumps(round(vals[last_idx], 2) if vals[last_idx] is not None else 0)}],
            symbol:'circle',symbolSize:8,
            itemStyle:{{color:dark?'{color_dark}':'{color_light}',borderColor:'#fff',borderWidth:2}},
            label:{{show:false}}
          }}]
        }},"""

        series_js_parts.append(f"""
      {{
        name:'{year}',type:'line',
        data:{vals_json},
        symbol:'none',
        connectNulls:false,
        lineStyle:{{width:{width},color:dark?'{color_dark}':'{color_light}',type:{line_type},opacity:{opacity}}},
        itemStyle:{{color:dark?'{color_dark}':'{color_light}'}},
        emphasis:{{lineStyle:{{width:{width + 1}}}}},{mark_point_js}
        z:{10 - i}
      }}""")

    return ",".join(series_js_parts)


def build_year_line_chart_js(year_data_list, y_axis_name="", y_formatter="null", value_suffix=""):
    """
    Genera el JS completo para un chart year-by-line.
    year_data_list: [(year, [12 vals]), ...] ordered newest-first
    """
    categories = json.dumps(MESES_CORTO)
    series_js = build_year_line_series_js(year_data_list, value_suffix)

    # Build legend data
    legend_data = json.dumps([str(y) for y, _ in year_data_list])

    # Build tooltip formatter with year colors
    tooltip_items = []
    for i, (year, _) in enumerate(year_data_list):
        is_covid = (year == COVID_YEAR)
        color_idx = min(i, len(YEAR_COLORS) - 1)
        c_light = COVID_COLOR if is_covid else YEAR_COLORS[color_idx]
        c_dark = COVID_COLOR_DARK if is_covid else YEAR_COLORS_DARK[color_idx]
        tooltip_items.append(f"'{year}':['{c_light}','{c_dark}']")

    y_formatter_js = ""
    if y_formatter == "thousands":
        y_formatter_js = "formatter:function(v){return (v/1000).toFixed(0)+'k'}"
    elif y_formatter == "integer":
        y_formatter_js = "formatter:function(v){return Math.round(v).toLocaleString('es-BO')}"
    elif y_formatter == "decimal1":
        y_formatter_js = "formatter:function(v){return v.toFixed(1)}"

    return f"""
    var dark=dk();
    var yColors={{{','.join(tooltip_items)}}};
    var option={{
      backgroundColor:'transparent',
      tooltip:{{
        trigger:'axis',
        backgroundColor:dark?'#080808':'rgba(13,27,42,.96)',
        borderColor:dark?'#2A3A50':'#1A2940',
        textStyle:{{color:'#E2E8F0',fontSize:12,fontFamily:'Inter'}},
        extraCssText:'border-radius:10px;box-shadow:0 8px 32px rgba(0,0,0,.3);',
        formatter:function(params){{
          var month=params[0].axisValue;
          var s='<div style="color:#E9D8A6;font-weight:700;margin-bottom:6px;font-family:Playfair Display,serif">'+month+'</div>';
          params.sort(function(a,b){{return parseInt(b.seriesName)-parseInt(a.seriesName)}});
          params.forEach(function(p){{
            if(p.value!=null){{
              var c=yColors[p.seriesName];
              var color=c?(dark?c[1]:c[0]):p.color;
              s+='<div style="display:flex;align-items:center;gap:6px;margin:2px 0">';
              s+='<span style="display:inline-block;width:10px;height:3px;border-radius:1px;background:'+color+'"></span>';
              s+='<span style="color:#94A3B8;font-size:11px;min-width:32px">'+p.seriesName+'</span>';
              s+='<span style="font-family:JetBrains Mono;font-weight:600;color:#fff;font-size:12px">'+p.value.toLocaleString('es-BO',{{maximumFractionDigits:1}})+'{value_suffix}</span>';
              s+='</div>';
            }}
          }});
          return s;
        }}
      }},
      legend:{{
        data:{legend_data},
        bottom:8,
        textStyle:{{color:dark?'#64748B':'#94A3B8',fontSize:11,fontFamily:'JetBrains Mono',fontWeight:500}},
        itemWidth:18,itemHeight:3,itemGap:14,
        type:'scroll'
      }},
      grid:{{left:65,right:20,top:20,bottom:55}},
      xAxis:{{
        type:'category',
        data:{categories},
        axisLine:{{lineStyle:{{color:dark?'#2A3A50':'#E2DDD3'}}}},
        axisTick:{{show:false}},
        axisLabel:{{fontFamily:'Inter',fontSize:10,fontWeight:500,color:dark?'#64748B':'#94A3B8'}}
      }},
      yAxis:{{
        type:'value',
        name:'{y_axis_name}',
        nameTextStyle:{{fontFamily:'Inter',fontSize:10,color:dark?'#64748B':'#94A3B8',padding:[0,0,0,0]}},
        splitLine:{{lineStyle:{{color:dark?'#1E1E1E':'#F1F5F9',type:'dashed'}}}},
        axisLine:{{show:false}},
        axisTick:{{show:false}},
        axisLabel:{{fontFamily:'JetBrains Mono',fontSize:11,color:dark?'#64748B':'#94A3B8'{(',' + y_formatter_js) if y_formatter_js else ''}}}
      }},
      series:[{series_js}
      ]
    }};"""


# -- Chart builders -------------------------------------------------------

def write_chart(filename, title, subtitle, chart_js, accent="#0D9488"):
    """Escribe archivo HTML."""
    html = _html_template(title, subtitle, chart_js, accent)
    out = EMBED / filename
    out.write_text(html, encoding="utf-8")
    print(f"  -> {out}")


# Chart 9: Cemento -- YEAR-BY-LINE
def build_cemento():
    print("[9] Produccion de cemento (year-by-line)...")
    filepath = DATOS / "cemento" / "produccion_cemento_depto_1991_2026.xlsx"
    data = parse_monthly_xlsx(filepath, "C1", col_periodo=3, col_map={"TOTAL": 10})
    total = data["TOTAL"]

    # Select last 7 years available
    all_years = sorted(total.keys())
    years = all_years[-7:]  # e.g. 2020-2026
    year_rows = build_year_data(total, years)
    year_rows.reverse()  # newest first

    chart_js = build_year_line_chart_js(
        year_rows,
        y_axis_name="Toneladas metricas",
        y_formatter="thousands",
        value_suffix=" TM"
    )

    write_chart(
        "cemento_produccion.html",
        "Producción de cemento",
        "Toneladas métricas, comparación mensual por año",
        chart_js
    )


# Chart 10: Permisos -- YEAR-BY-LINE
def build_permisos():
    print("[10] Permisos de construccion (year-by-line)...")
    filepath = DATOS / "construccion" / "numero_permisos_2008_2026.xlsx"
    data = parse_monthly_xls(filepath, sheet_index=0, col_periodo=1, col_map={"TOTAL": 2})
    total = data["TOTAL"]

    all_years = sorted(total.keys())
    years = all_years[-7:]
    year_rows = build_year_data(total, years)
    year_rows.reverse()

    chart_js = build_year_line_chart_js(
        year_rows,
        y_axis_name="N° registros",
        y_formatter="integer",
        value_suffix=""
    )

    write_chart(
        "permisos_construccion.html",
        "Permisos de construcción",
        "Número de registros, comparación mensual por año",
        chart_js
    )


# Chart 11: Manufactura -- DUAL AXIS (quarterly, standard time series)
def build_manufactura():
    print("[11] Manufactura: produccion + capacidad (dual axis)...")
    fp_prod = DATOS / "manufactura" / "indice_produccion_manufactura_2017_2025.xlsx"
    labels_p, series_p = parse_quarterly_xlsx(fp_prod, "Indice", col_periodo=2, col_map={"Produccion": 3})

    fp_cap = DATOS / "manufactura" / "utilizacion_capacidad_instalada_2017_2025.xlsx"
    labels_c, series_c = parse_quarterly_xlsx(fp_cap, "cuadro_1", col_periodo=2, col_map={"Capacidad": 3})

    # Align by common labels
    common = [l for l in labels_p if l in labels_c]
    prod_vals = []
    cap_vals = []
    for l in common:
        prod_vals.append(series_p["Produccion"][labels_p.index(l)])
        cap_vals.append(series_c["Capacidad"][labels_c.index(l)])

    labels_json = json.dumps(common)
    prod_json = json.dumps(json_safe(prod_vals))
    cap_json = json.dumps(json_safe(cap_vals))

    chart_js = f"""
    var dark=dk();
    var option={{
      backgroundColor:'transparent',
      tooltip:{{
        trigger:'axis',
        backgroundColor:dark?'#080808':'rgba(13,27,42,.96)',
        borderColor:dark?'#2A3A50':'#1A2940',
        textStyle:{{color:'#E2E8F0',fontSize:12,fontFamily:'Inter'}},
        extraCssText:'border-radius:10px;box-shadow:0 8px 32px rgba(0,0,0,.3);',
        formatter:function(params){{
          var s='<div style="color:#E9D8A6;font-weight:700;margin-bottom:6px;font-family:Playfair Display,serif">'+params[0].axisValue+'</div>';
          params.forEach(function(p){{
            if(p.value!=null){{
              var unit=p.seriesIndex===0?' pts':' %';
              s+='<div style="display:flex;align-items:center;gap:6px;margin:2px 0">';
              s+='<span style="display:inline-block;width:10px;height:3px;border-radius:1px;background:'+p.color+'"></span>';
              s+='<span style="color:#94A3B8;font-size:11px">'+p.seriesName+'</span>';
              s+='<span style="font-family:JetBrains Mono;font-weight:600;color:#fff;font-size:12px">'+p.value.toFixed(1)+unit+'</span>';
              s+='</div>';
            }}
          }});
          return s;
        }}
      }},
      legend:{{
        data:['\\u00cdndice de producci\\u00f3n','Utilizaci\\u00f3n capacidad (%)'],
        bottom:8,
        textStyle:{{color:dark?'#64748B':'#94A3B8',fontSize:11,fontFamily:'Inter',fontWeight:500}},
        itemWidth:18,itemHeight:3,itemGap:14
      }},
      grid:{{left:60,right:60,top:20,bottom:50}},
      xAxis:{{
        type:'category',data:{labels_json},
        axisLine:{{lineStyle:{{color:dark?'#2A3A50':'#E2DDD3'}}}},
        axisTick:{{show:false}},
        axisLabel:{{fontFamily:'JetBrains Mono',fontSize:10,fontWeight:500,color:dark?'#64748B':'#94A3B8',rotate:45}}
      }},
      yAxis:[
        {{
          type:'value',name:'\\u00cdndice (base 2017)',
          nameTextStyle:{{fontFamily:'Inter',fontSize:10,color:'#0D9488'}},
          axisLabel:{{fontFamily:'JetBrains Mono',fontSize:11,color:'#0D9488'}},
          axisLine:{{show:true,lineStyle:{{color:'#0D9488'}}}},
          splitLine:{{lineStyle:{{color:dark?'#1E1E1E':'#F1F5F9',type:'dashed'}}}},
          axisTick:{{show:false}}
        }},
        {{
          type:'value',name:'Capacidad (%)',
          nameTextStyle:{{fontFamily:'Inter',fontSize:10,color:'#D97706'}},
          axisLabel:{{fontFamily:'JetBrains Mono',fontSize:11,color:'#D97706',formatter:'{{value}}%'}},
          axisLine:{{show:true,lineStyle:{{color:'#D97706'}}}},
          splitLine:{{show:false}},
          axisTick:{{show:false}}
        }}
      ],
      series:[
        {{
          name:'\\u00cdndice de producci\\u00f3n',type:'line',data:{prod_json},
          yAxisIndex:0,symbol:'circle',symbolSize:5,
          lineStyle:{{width:2.5,color:'#0D9488'}},
          itemStyle:{{color:'#0D9488'}},
          areaStyle:{{color:new echarts.graphic.LinearGradient(0,0,0,1,[
            {{offset:0,color:dark?'rgba(13,148,136,.2)':'rgba(13,148,136,.1)'}},
            {{offset:1,color:'rgba(13,148,136,0)'}}
          ])}}
        }},
        {{
          name:'Utilizaci\\u00f3n capacidad (%)',type:'line',data:{cap_json},
          yAxisIndex:1,symbol:'diamond',symbolSize:5,
          lineStyle:{{width:2.5,color:'#D97706'}},
          itemStyle:{{color:'#D97706'}}
        }}
      ]
    }};"""

    write_chart(
        "manufactura_produccion_capacidad.html",
        "Industria manufacturera: producción y capacidad instalada",
        "Trimestral 2017-2025 · Base 2017=100 (producción) / % (capacidad)",
        chart_js,
        accent="#0D9488"
    )


# Chart 12: Energia -- YEAR-BY-LINE
def build_energia():
    print("[12] Consumo de energia electrica (year-by-line)...")
    filepath = DATOS / "energia" / "indice_consumo_energia_electrica_2017_2026.xlsx"
    # Sheet name has encoding issues, use first sheet via fallback
    data = parse_monthly_xlsx(filepath, "Indice", col_periodo=2, col_map={"INDICE GENERAL": 9})
    total = data["INDICE GENERAL"]

    all_years = sorted(total.keys())
    years = all_years[-7:]
    year_rows = build_year_data(total, years)
    year_rows.reverse()

    chart_js = build_year_line_chart_js(
        year_rows,
        y_axis_name="Índice (2017=100)",
        y_formatter="decimal1",
        value_suffix=""
    )

    write_chart(
        "energia_consumo.html",
        "Consumo de energía eléctrica",
        "Índice general, comparación mensual por año (base 2017=100)",
        chart_js
    )


# Chart 13: Transporte -- YEAR-BY-LINE
def build_transporte():
    print("[13] Transporte indice general (year-by-line)...")
    filepath = DATOS / "transporte" / "indice_general_transporte_2017_2026.xlsx"
    data = parse_monthly_xlsx(filepath, "Indice", col_periodo=2, col_map={"INDICE": 8})
    total = data["INDICE"]

    all_years = sorted(total.keys())
    years = all_years[-7:]
    year_rows = build_year_data(total, years)
    year_rows.reverse()

    chart_js = build_year_line_chart_js(
        year_rows,
        y_axis_name="Índice (2017=100)",
        y_formatter="decimal1",
        value_suffix=""
    )

    write_chart(
        "transporte_indice.html",
        "Índice de transporte",
        "Comparación mensual por año (base 2017=100)",
        chart_js
    )


# Chart 14: Carga ferroviaria -- YEAR-BY-LINE
def build_carga_ferroviaria():
    print("[14] Carga ferroviaria total (year-by-line)...")
    filepath = DATOS / "transporte" / "flujo_ferroviario_1999_2026.xlsx"
    data = parse_monthly_xlsx(
        filepath, "Flujo", col_periodo=2,
        col_map={"Red Andina Carga": 4, "Red Oriental Carga": 7},
        start_year=2000
    )

    andina = data["Red Andina Carga"]
    oriental = data["Red Oriental Carga"]

    # Sum both into total cargo
    all_years_set = set(list(andina.keys()) + list(oriental.keys()))
    total = {}
    for y in all_years_set:
        total[y] = {}
        for m in range(1, 13):
            a = andina.get(y, {}).get(m)
            o = oriental.get(y, {}).get(m)
            if a is not None and o is not None:
                total[y][m] = a + o
            elif a is not None:
                total[y][m] = a
            elif o is not None:
                total[y][m] = o

    all_years = sorted(total.keys())
    years = all_years[-7:]
    year_rows = build_year_data(total, years)
    year_rows.reverse()

    chart_js = build_year_line_chart_js(
        year_rows,
        y_axis_name="Toneladas metricas",
        y_formatter="thousands",
        value_suffix=" TM"
    )

    write_chart(
        "transporte_carga_ferroviaria.html",
        "Carga ferroviaria total",
        "Toneladas métricas, comparación mensual por año",
        chart_js
    )


# -- Main -----------------------------------------------------------------

def main():
    print("=" * 60)
    print("Generando charts de indicadores adelantados (POPULI DS)")
    print("=" * 60)

    build_cemento()
    build_permisos()
    build_manufactura()
    build_energia()
    build_transporte()
    build_carga_ferroviaria()

    print()
    print("=" * 60)
    print(f"Listo: 6 archivos HTML en {EMBED}")
    print("=" * 60)


if __name__ == "__main__":
    main()
