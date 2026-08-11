"""
build_pib_charts.py  --  POPULI Design System
Reads INE Bolivia Excel files and generates 8 standalone ECharts HTML embeds.
Output: ../embed/

Each file follows the POPULI design system:
  - CSS custom properties (light/dark)
  - Google Fonts: Inter, Playfair Display, JetBrains Mono
  - Theme switching via ?theme=dark or postMessage
  - Watermark + source footer
"""

import json
import math
import re
from pathlib import Path
from typing import Optional

import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
DATOS_DIR = BASE_DIR / "datos"
EMBED_DIR = BASE_DIR / "embed"
EMBED_DIR.mkdir(parents=True, exist_ok=True)

# Base 2017 files
FILE_PIB_NIVEL = DATOS_DIR / "pib_trimestral_base2017/encadenadas/07_pib_actividad_encadenadas.xlsx"
FILE_PIB_VARIACION = DATOS_DIR / "pib_trimestral_base2017/encadenadas/09_pib_actividad_variacion_encadenadas.xlsx"
FILE_PIB_CONTRIBUCION = DATOS_DIR / "pib_trimestral_base2017/encadenadas/11_pib_actividad_contribucion_encadenadas.xlsx"
FILE_PIB_ESTRUCTURA = DATOS_DIR / "pib_trimestral_base2017/corrientes/02_pib_actividad_estructura_corrientes.xlsx"
FILE_GASTO_NIVEL = DATOS_DIR / "pib_trimestral_base2017/encadenadas/19_pib_gasto_encadenadas.xlsx"
FILE_GASTO_CRECIMIENTO = DATOS_DIR / "pib_trimestral_base2017/encadenadas/20_pib_gasto_crecimiento_encadenadas.xlsx"
FILE_GASTO_CONTRIBUCION = DATOS_DIR / "pib_trimestral_base2017/encadenadas/22_pib_gasto_contribucion_encadenadas.xlsx"

# Base 1990
FILE_BASE1990_VARIACION = DATOS_DIR / "pib_trimestral_base1990/actividad/04_pib_actividad_variacion_similar.xlsx"

# ── Sector palette & names ──────────────────────────────────────────────────
# Paleta oficial POPULI — ver Proyectos/populi-marca/paleta.py (fuente única).
# 11 sectores > las 4 series validadas, así que va CATEGORICAL_12 (4 familias × 3
# pasos). Admin. Pública queda en gris a propósito: no es un sector productivo.
SECTOR_COLORS = {
    "agro":    "#0A9396",
    "extract": "#A86E00",
    "manuf":   "#005F73",
    "elec":    "#EE9B00",
    "const":   "#C71E1D",
    "comer":   "#94D2BD",
    "trans":   "#9B2226",
    "aloja":   "#DF5D25",
    "finan":   "#E9D8A6",
    "admin":   "#5C6B70",
    "comun":   "#E8706B",
}
SECTOR_COLORS_LIST = list(SECTOR_COLORS.values())

SECTOR_SHORT = [
    "Agropecuaria",
    "Extractiva",
    "Manufactura",
    "Electricidad",
    "Construcción",
    "Comercio",
    "Transporte",
    "Alojamiento",
    "Financieras",
    "Admin. Pública",
    "Comunales",
]

GASTO_NAMES = [
    "Consumo Gobierno",
    "Consumo Hogares",
    "FBK",
    "FBCF",
    "Exportaciones",
    "Importaciones",
]
# Los pares comparten familia para que la relación se lea en el color:
# consumo (petróleo claro/oscuro) · inversión FBK-FBCF (ámbar) · comercio exterior
# usa el par validado #0A9396/#C71E1D, que es el que mejor contrasta entre sí.
GASTO_COLORS = {
    "Consumo Gobierno": "#005F73",
    "Consumo Hogares": "#94D2BD",
    "FBK": "#EE9B00",
    "FBCF": "#A86E00",
    "Exportaciones": "#0A9396",
    "Importaciones": "#C71E1D",
}


# ── Parsing helpers ──────────────────────────────────────────────────────────

def _safe_round(v, n=4):
    if v is not None and isinstance(v, (int, float)) and not math.isnan(v):
        return round(v, n)
    return None


def parse_base2017_transposed(filepath, sheet_name=None, data_rows=None):
    """
    Parse base-2017 transposed layout.
    Returns (quarter_labels, {row_label: [values...]})
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    max_col = ws.max_column
    year_row = [ws.cell(10, c).value for c in range(1, max_col + 1)]
    q_row = [ws.cell(11, c).value for c in range(1, max_col + 1)]

    current_year = None
    years_filled = []
    for v in year_row:
        if v is not None and v not in ("", " "):
            s = str(v).strip()
            m = re.match(r"(\d{4})", s)
            if m:
                current_year = int(m.group(1))
        years_filled.append(current_year)

    quarter_labels = []
    data_col_indices = []
    qmap = {"I": "T1", "II": "T2", "III": "T3", "IV": "T4"}
    for ci in range(len(q_row)):
        qv = q_row[ci]
        yr = years_filled[ci]
        if qv is not None and yr is not None:
            qs = str(qv).strip()
            if qs in qmap:
                quarter_labels.append(f"{yr}-{qmap[qs]}")
                data_col_indices.append(ci + 1)

    result = {}
    if data_rows is None:
        data_rows = list(range(13, 28))
    for r in data_rows:
        name_cell = ws.cell(r, 2).value
        if name_cell is None:
            continue
        name = str(name_cell).strip()
        values = [_safe_round(ws.cell(r, c).value) for c in data_col_indices]
        result[name] = values

    wb.close()
    return quarter_labels, result


def parse_base1990_quarterly(filepath, sheet_name="01.01.04", pib_col=16, start_year=2000):
    """
    Parse base-1990 file with periods in ROWS.
    Returns (quarter_labels, values) for quarterly rows only.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb[sheet_name]

    quarter_labels = []
    values = []
    current_year = None

    # Regex: match roman numeral quarter at start (after optional spaces)
    q_re = re.compile(r"^\s*(IV|III|II|I)\s+Trimestre", re.IGNORECASE)
    roman_to_q = {"I": "T1", "II": "T2", "III": "T3", "IV": "T4"}

    for r in range(12, ws.max_row + 1):
        cell_a = ws.cell(r, 1).value
        if cell_a is None:
            continue
        s = str(cell_a).strip()

        m_year = re.match(r"^(\d{4})", s)
        if m_year:
            current_year = int(m_year.group(1))
            continue

        m_q = q_re.match(s)
        if m_q:
            roman = m_q.group(1).upper()
            q_label = roman_to_q[roman]
            if current_year and current_year >= start_year:
                v = ws.cell(r, pib_col).value
                quarter_labels.append(f"{current_year}-{q_label}")
                values.append(_safe_round(v))

    wb.close()
    return quarter_labels, values


def json_safe(values, decimals=2):
    return [round(v, decimals) if v is not None else None for v in values]


# ── Helpers to extract sector / gasto keys from parsed data ─────────────────

def split_sector_data(data_dict):
    """
    From a parsed dict, separate PIB total (row 13) from sector rows (17-27).
    Returns (pib_total_key, pib_total_values, [(sector_key, values), ...])
    """
    keys = list(data_dict.keys())
    pib_total_key = keys[0]
    pib_total = data_dict[pib_total_key]

    sectors = []
    for k in keys:
        ku = k.upper()
        if "PRODUCTO INTERNO" in ku or "DERECHOS" in ku or "VALOR AGREGADO" in ku or ku == "":
            continue
        sectors.append((k, data_dict[k]))

    return pib_total_key, pib_total, sectors[:11]


def classify_gasto(k):
    """Return (display_name, color) for an expenditure key, or None to skip."""
    ku = k.upper()
    if "PRODUCTO" in ku:
        return ("PIB Total", "#0A9396")
    if "ADMINISTRACI" in ku:
        return ("Consumo Gobierno", "#005F73")
    if "HOGARES" in ku:
        return ("Consumo Hogares", "#0A9396")
    if "FORMACI" in ku and "FIJO" not in ku:
        return ("FBK", "#EE9B00")
    if "FIJO" in ku:
        return ("FBCF", "#A86E00")
    if "EXPORTACION" in ku:
        return ("Exportaciones", "#0A9396")
    if "IMPORTACION" in ku:
        return ("Importaciones", "#C71E1D")
    return None


# ── HTML Template (POPULI Design System) ────────────────────────────────────

def populi_html(title, chart_height, section_title, section_subtitle,
                accent_bar_color, build_chart_js, source_label="INE"):
    """
    Generate a standalone POPULI-themed HTML embed.
    build_chart_js: JS function body that calls chart.setOption(...)
                    It can reference dk() and the chart instance.
    """
    return f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>{title}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300..800&family=Playfair+Display:ital,wght@0,400..900;1,400..900&family=JetBrains+Mono:wght@400;500;600;700&display=swap" rel="stylesheet"/>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
:root{{
  --populi:#C71E1D;--populi-dark:#8B1A1A;--populi-light:#E8706B;--populi-gold:#EE9B00;
  --cream:#F5EFE0;--brown:#5C3D1E;--brown-dark:#3D2B1F;
  --navy:#0D1B2A;--navy-light:#1A2940;
  --slate:#475569;--slate-light:#64748B;
  --warm-white:#FAF8F3;--light-gray:#F1EDE5;
  --border:#E2DDD3;--dark-border:#2A3A50;
  --bg:var(--warm-white);--card:#FFFFFF;--text:var(--brown-dark);--muted:var(--slate-light);
  --shadow:0 1px 3px rgba(13,27,42,.04),0 4px 16px rgba(13,27,42,.06);
  --radius:12px;
  --ct-pib:#0A9396;--ct-agro:#0A9396;--ct-extract:#EE9B00;--ct-manuf:#005F73;
  --ct-elec:#9B2226;--ct-const:#C71E1D;--ct-comer:#94D2BD;--ct-trans:#005F73;
  --ct-aloja:#DF5D25;--ct-finan:#EE9B00;--ct-admin:#64748B;--ct-comun:#0A9396;
}}
[data-theme="dark"]{{
  --bg:#080808;--card:#141414;--text:#E2E8F0;--muted:#64748B;
  --border:#2A3A50;--light-gray:#1A1A1A;--cream:#1A2940;
}}
body{{
  font-family:'Inter',system-ui,sans-serif;
  background:var(--bg);color:var(--text);
  -webkit-font-smoothing:antialiased;
  transition:background .25s,color .25s;
}}
.embed-container{{max-width:1200px;margin:0 auto;padding:20px 16px}}
.section-title{{
  font-family:'Playfair Display',Georgia,serif;font-size:1.55rem;font-weight:700;
  color:var(--navy);display:flex;align-items:center;gap:10px;
}}
[data-theme="dark"] .section-title{{color:#fff}}
.accent-bar{{width:4px;height:26px;border-radius:2px;flex-shrink:0}}
.section-subtitle{{font-size:.78rem;color:var(--muted);padding-left:14px;margin-top:2px}}
.chart-card{{
  background:var(--card);border:1px solid var(--border);border-radius:var(--radius);
  padding:16px;display:flex;flex-direction:column;overflow:hidden;margin-top:12px;
}}
[data-theme="dark"] .chart-card{{background:#141414;border-color:var(--dark-border)}}
#chart{{width:100%;height:{chart_height}px}}
.chart-source{{
  display:flex;align-items:center;justify-content:space-between;margin-top:8px;padding:0 2px;
}}
.chart-source-text{{font-size:.58rem;color:var(--muted);opacity:.6}}
.chart-source-text a{{color:var(--populi);font-weight:600;text-decoration:none}}
.watermark{{height:15px;width:auto;opacity:.18}}
[data-theme="dark"] .watermark{{opacity:.10}}
</style>
</head>
<body>
<div class="embed-container">
  <div class="section-title">
    <span class="accent-bar" style="background:{accent_bar_color}"></span>
    {section_title}
  </div>
  <div class="section-subtitle">{section_subtitle}</div>
  <div class="chart-card">
    <div id="chart"></div>
    <div class="chart-source">
      <span class="chart-source-text">Fuente: <a>{source_label}</a> &middot; Elaboraci&oacute;n: <a>Centro de Estudios POPULI</a></span>
      <svg viewBox="0 0 310 130" class="watermark"><text x="6" y="112" font-family="'Playfair Display'" font-size="145" font-weight="700" fill="#C71E1D">P</text><text x="92" y="87" font-family="'Playfair Display'" font-size="55" font-weight="400" font-style="italic" fill="#3D2B1F">opuli</text></svg>
    </div>
  </div>
</div>
<script>
(function(){{
  /* ── Theme ─────────────────────────────────────── */
  var params=new URLSearchParams(window.location.search);
  if(params.get('theme')==='dark') document.documentElement.dataset.theme='dark';
  window.addEventListener('message',function(e){{if(e.data&&e.data.theme) document.documentElement.dataset.theme=e.data.theme}});
  function dk(){{return document.documentElement.dataset.theme==='dark'}}

  /* ── Chart ─────────────────────────────────────── */
  var chartDom=document.getElementById('chart');
  var chart=echarts.init(chartDom);

  function buildChart(){{
{build_chart_js}
  }}

  buildChart();
  new MutationObserver(function(){{buildChart()}}).observe(document.documentElement,{{attributes:true,attributeFilter:['data-theme']}});
  window.addEventListener('resize',function(){{chart.resize()}});
}})();
</script>
</body>
</html>'''


# ── Reusable ECharts option fragments ───────────────────────────────────────

def js_tooltip(unit="", is_pp=False):
    """Generate POPULI tooltip config as JS string."""
    fmt_val = "pp" if is_pp else unit
    return f"""{{
      trigger:'axis',
      backgroundColor:dk()?'#080808':'rgba(13,27,42,.96)',
      borderColor:dk()?'#2A3A50':'#1A2940',
      borderWidth:1,
      textStyle:{{fontFamily:'Inter',fontSize:12,color:'#fff'}},
      padding:[12,16],
      extraCssText:'border-radius:10px;box-shadow:0 8px 32px rgba(0,0,0,.3);',
      formatter:function(params){{
        if(!params||!params.length) return '';
        var s='<div style="color:#E9D8A6;font-weight:600;margin-bottom:6px">'+params[0].axisValue+'</div>';
        params.forEach(function(p){{
          if(p.value!=null){{
            s+=p.marker+' <span style="color:#ccc">'+p.seriesName+'</span> '
              +'<span style="font-family:\\'JetBrains Mono\\';font-weight:700;color:#fff">'
              +p.value.toFixed(2)+' {fmt_val}</span><br/>';
          }}
        }});
        return s;
      }}
    }}"""


def js_tooltip_bar_stack(unit="pp"):
    return f"""{{
      trigger:'axis',
      axisPointer:{{type:'shadow'}},
      backgroundColor:dk()?'#080808':'rgba(13,27,42,.96)',
      borderColor:dk()?'#2A3A50':'#1A2940',
      borderWidth:1,
      textStyle:{{fontFamily:'Inter',fontSize:12,color:'#fff'}},
      padding:[12,16],
      extraCssText:'border-radius:10px;box-shadow:0 8px 32px rgba(0,0,0,.3);',
      formatter:function(params){{
        if(!params||!params.length) return '';
        var s='<div style="color:#E9D8A6;font-weight:600;margin-bottom:6px">'+params[0].axisValue+'</div>';
        params.forEach(function(p){{
          if(p.value!=null){{
            s+=p.marker+' <span style="color:#ccc">'+p.seriesName+'</span> '
              +'<span style="font-family:\\'JetBrains Mono\\';font-weight:700;color:#fff">'
              +p.value.toFixed(2)+' {unit}</span><br/>';
          }}
        }});
        return s;
      }}
    }}"""


def js_xaxis(labels_var="labels", rotate=45, interval="'auto'"):
    return f"""{{
      type:'category',
      data:{labels_var},
      boundaryGap:false,
      axisLine:{{lineStyle:{{color:dk()?'#2A3A50':'#E2DDD3'}}}},
      axisTick:{{show:false}},
      axisLabel:{{
        fontFamily:'Inter',fontSize:10,fontWeight:500,
        color:dk()?'#64748B':'#64748B',
        rotate:{rotate},
        interval:{interval}
      }}
    }}"""


def js_xaxis_bar(labels_var="labels", rotate=45, interval="'auto'"):
    return f"""{{
      type:'category',
      data:{labels_var},
      axisLine:{{lineStyle:{{color:dk()?'#2A3A50':'#E2DDD3'}}}},
      axisTick:{{show:false}},
      axisLabel:{{
        fontFamily:'Inter',fontSize:10,fontWeight:500,
        color:dk()?'#64748B':'#64748B',
        rotate:{rotate},
        interval:{interval}
      }}
    }}"""


def js_yaxis(name="", fmt="{value}", left_pad=52):
    return f"""{{
      type:'value',
      name:'{name}',
      nameTextStyle:{{fontFamily:'Inter',fontSize:11,color:dk()?'#64748B':'#64748B',padding:[0,0,0,0]}},
      splitLine:{{lineStyle:{{color:dk()?'#141414':'#F1F5F9',type:'dashed'}}}},
      axisLine:{{show:false}},
      axisTick:{{show:false}},
      axisLabel:{{
        fontFamily:'JetBrains Mono',fontSize:11,
        color:dk()?'#64748B':'#64748B',
        formatter:'{fmt}'
      }}
    }}"""


def js_line_series(name, data_var, color, width=2.5, dashed=False, z=2):
    """Line series with area gradient, markPoint on last valid data point."""
    dash_str = "type:'dashed'," if dashed else ""
    return f"""{{
      name:'{name}',type:'line',data:{data_var},
      symbol:'none',z:{z},
      lineStyle:{{width:{width},color:'{color}',{dash_str}}},
      itemStyle:{{color:'{color}'}},
      areaStyle:{{color:new echarts.graphic.LinearGradient(0,0,0,1,[
        {{offset:0,color:'{color}18'}},
        {{offset:1,color:'{color}02'}}
      ])}},
      markPoint:{{
        data:[{{
          type:'max',
          symbol:'circle',symbolSize:7,
          coord:(function(){{
            var d={data_var};
            for(var i=d.length-1;i>=0;i--){{if(d[i]!=null)return[i,d[i]]}}
            return[0,0];
          }})(),
          itemStyle:{{color:'{color}',borderColor:dk()?'#141414':'#fff',borderWidth:2}},
          label:{{show:false}}
        }}],
        symbol:'circle',symbolSize:7,
        animation:false
      }}
    }}"""


def js_line_series_plain(name, data_var, color, width=2.5, dashed=False, z=2):
    """Line series without area gradient."""
    dash_str = "type:'dashed'," if dashed else ""
    return f"""{{
      name:'{name}',type:'line',data:{data_var},
      symbol:'none',z:{z},
      lineStyle:{{width:{width},color:'{color}',{dash_str}}},
      itemStyle:{{color:'{color}'}}
    }}"""


# ── Chart Builders ───────────────────────────────────────────────────────────

def build_chart1():
    """Chart 1: Long GDP growth series (base 1990 + base 2017 overlap)"""
    print("  [1] pib_crecimiento_serie_larga.html")

    labels_90, vals_90 = parse_base1990_quarterly(FILE_BASE1990_VARIACION, start_year=2000)

    labels_17, data_17 = parse_base2017_transposed(FILE_PIB_VARIACION, data_rows=[13])
    pib_key = list(data_17.keys())[0]
    vals_17 = data_17[pib_key]

    all_labels = sorted(set(labels_90 + labels_17),
                        key=lambda x: (int(x.split("-")[0]), int(x.split("-T")[1])))

    map_90 = dict(zip(labels_90, vals_90))
    map_17 = dict(zip(labels_17, vals_17))

    d90 = json.dumps(json_safe([map_90.get(l) for l in all_labels]))
    d17 = json.dumps(json_safe([map_17.get(l) for l in all_labels]))
    lbl = json.dumps(all_labels)

    covid_idx = all_labels.index("2020-T2") if "2020-T2" in all_labels else -1

    chart_js = f"""
    var labels={lbl};
    var d90={d90};
    var d17={d17};
    chart.setOption({{
      backgroundColor:'transparent',
      grid:{{top:30,right:16,bottom:44,left:52}},
      tooltip:{js_tooltip('%')},
      legend:{{
        data:['Base 1990','Base 2017'],
        bottom:4,
        textStyle:{{fontFamily:'Inter',fontSize:11,color:dk()?'#64748B':'#64748B'}}
      }},
      xAxis:{js_xaxis('labels',45,3)},
      yAxis:{js_yaxis('%','{value}%')},
      series:[
        {{
          name:'Base 1990',type:'line',data:d90,
          symbol:'none',z:1,
          lineStyle:{{width:1.8,color:'#9CA3AF'}},
          itemStyle:{{color:'#9CA3AF'}},
          areaStyle:{{color:new echarts.graphic.LinearGradient(0,0,0,1,[
            {{offset:0,color:'#9CA3AF18'}},
            {{offset:1,color:'#9CA3AF02'}}
          ])}},
          connectNulls:false
        }},
        {{
          name:'Base 2017',type:'line',data:d17,
          symbol:'none',z:3,
          lineStyle:{{width:2.5,color:'#0A9396'}},
          itemStyle:{{color:'#0A9396'}},
          areaStyle:{{color:new echarts.graphic.LinearGradient(0,0,0,1,[
            {{offset:0,color:'#0A939618'}},
            {{offset:1,color:'#0A939602'}}
          ])}},
          connectNulls:false,
          markLine:{{
            silent:true,symbol:'none',
            lineStyle:{{color:dk()?'#475569':'#64748B',type:'solid',width:1}},
            data:[{{yAxis:0}}],
            label:{{show:false}}
          }},
          markPoint:{{
            data:[{{
              coord:['{all_labels[covid_idx] if covid_idx >= 0 else ""}',{json.dumps(json_safe([map_17.get(all_labels[covid_idx])])[0]) if covid_idx >= 0 else 'null'}],
              symbol:'pin',symbolSize:40,
              itemStyle:{{color:'#C71E1D'}},
              label:{{show:true,formatter:'COVID-19',fontSize:9,fontWeight:700,color:'#fff'}},
            }}],
            animation:false
          }}
        }}
      ]
    }},true);"""

    html = populi_html(
        title="PIB Bolivia - Serie larga",
        chart_height=500,
        section_title="Crecimiento del PIB de Bolivia",
        section_subtitle="Variación interanual trimestral (%). Base 1990 desde 2000-T1 y Base 2017 desde 2018-T1",
        accent_bar_color="#0A9396",
        build_chart_js=chart_js,
        source_label="INE",
    )
    (EMBED_DIR / "pib_crecimiento_serie_larga.html").write_text(html, encoding="utf-8")


def build_chart2():
    """Chart 2: PIB level in chained Bs (bar chart)"""
    print("  [2] pib_nivel_encadenado.html")

    labels, data = parse_base2017_transposed(FILE_PIB_NIVEL, data_rows=[13])
    pib_key = list(data.keys())[0]
    vals = json_safe(data[pib_key], 1)

    lbl = json.dumps(labels)
    dv = json.dumps(vals)

    chart_js = f"""
    var labels={lbl};
    var vals={dv};
    chart.setOption({{
      backgroundColor:'transparent',
      grid:{{top:30,right:16,bottom:44,left:68}},
      tooltip:{{
        trigger:'axis',
        backgroundColor:dk()?'#080808':'rgba(13,27,42,.96)',
        borderColor:dk()?'#2A3A50':'#1A2940',
        borderWidth:1,
        textStyle:{{fontFamily:'Inter',fontSize:12,color:'#fff'}},
        padding:[12,16],
        extraCssText:'border-radius:10px;box-shadow:0 8px 32px rgba(0,0,0,.3);',
        formatter:function(params){{
          if(!params||!params.length) return '';
          var p=params[0];
          return '<div style="color:#E9D8A6;font-weight:600;margin-bottom:6px">'+p.axisValue+'</div>'
            +p.marker+' <span style="font-family:\\'JetBrains Mono\\';font-weight:700;color:#fff">'
            +p.value.toLocaleString('es-BO')+'</span> <span style="color:#ccc">MM Bs</span>';
        }}
      }},
      xAxis:{js_xaxis_bar('labels',45,1)},
      yAxis:{{
        type:'value',
        name:'Millones Bs encadenados',
        nameTextStyle:{{fontFamily:'Inter',fontSize:10,color:dk()?'#64748B':'#64748B'}},
        splitLine:{{lineStyle:{{color:dk()?'#141414':'#F1F5F9',type:'dashed'}}}},
        axisLine:{{show:false}},axisTick:{{show:false}},
        axisLabel:{{fontFamily:'JetBrains Mono',fontSize:11,color:dk()?'#64748B':'#64748B',
          formatter:function(v){{return (v/1000).toFixed(0)+'k'}}
        }},
        min:'dataMin'
      }},
      series:[{{
        type:'bar',data:vals,
        itemStyle:{{
          color:new echarts.graphic.LinearGradient(0,0,0,1,[
            {{offset:0,color:'#0A9396'}},
            {{offset:1,color:dk()?'rgba(10,147,150,.4)':'rgba(10,147,150,.6)'}}
          ]),
          borderRadius:[3,3,0,0]
        }},
        barMaxWidth:18
      }}]
    }},true);"""

    html = populi_html(
        title="PIB Nivel Encadenado",
        chart_height=440,
        section_title="PIB Trimestral Encadenado",
        section_subtitle="Millones de bolivianos encadenados, año de referencia 2017",
        accent_bar_color="#0A9396",
        build_chart_js=chart_js,
    )
    (EMBED_DIR / "pib_nivel_encadenado.html").write_text(html, encoding="utf-8")


def build_chart3():
    """Chart 3: Small multiples grid - GDP growth by sector"""
    print("  [3] pib_crecimiento_sectores.html")

    labels, data = parse_base2017_transposed(FILE_PIB_VARIACION, data_rows=list(range(13, 28)))
    _, pib_total, sectors = split_sector_data(data)

    lbl = json.dumps(labels)
    pib_js = json.dumps(json_safe(pib_total))

    n_cols = 4
    n_rows = 3
    panels = sectors[:11]
    names = SECTOR_SHORT[:11]
    colors = SECTOR_COLORS_LIST[:11]

    # If fewer than 12 cells, pad
    while len(panels) < n_cols * n_rows:
        panels.append(("", []))
        names.append("")
        colors.append("#999")

    # Build grid, xAxis, yAxis, series, title arrays as JS
    grids_js = []
    xaxes_js = []
    yaxes_js = []
    series_js = []
    titles_js = []

    pad_l, pad_r, pad_t, pad_b = 4, 2, 2, 5
    cw = (100 - pad_l - pad_r) / n_cols
    ch = (100 - pad_t - pad_b) / n_rows

    for idx in range(n_cols * n_rows):
        ri = idx // n_cols
        ci = idx % n_cols
        left = pad_l + ci * cw + 1.5
        top = pad_t + ri * ch + 3
        w = cw - 4
        h = ch - 7

        sk, sv = panels[idx] if idx < len(panels) else ("", [])
        sn = names[idx] if idx < len(names) else ""
        sc = colors[idx] if idx < len(colors) else "#999"

        grids_js.append(f"{{left:'{left:.1f}%',top:'{top:.1f}%',width:'{w:.1f}%',height:'{h:.1f}%'}}")

        xaxes_js.append(f"""{{
          type:'category',gridIndex:{idx},data:labels,
          axisLabel:{{show:false}},axisTick:{{show:false}},
          axisLine:{{lineStyle:{{color:dk()?'#2A3A50':'#E2DDD3'}}}},
          splitLine:{{show:false}}
        }}""")

        yaxes_js.append(f"""{{
          type:'value',gridIndex:{idx},
          axisLabel:{{fontFamily:'JetBrains Mono',fontSize:9,color:dk()?'#64748B':'#64748B'}},
          splitLine:{{lineStyle:{{color:dk()?'#141414':'#F1F5F9',type:'dashed'}}}},
          axisLine:{{show:false}},axisTick:{{show:false}}
        }}""")

        dvar = json.dumps(json_safe(sv))
        # Sector line
        series_js.append(f"""{{
          type:'line',xAxisIndex:{idx},yAxisIndex:{idx},data:{dvar},
          symbol:'none',lineStyle:{{color:'{sc}',width:2}},
          itemStyle:{{color:'{sc}'}},
          areaStyle:{{color:new echarts.graphic.LinearGradient(0,0,0,1,[
            {{offset:0,color:'{sc}18'}},{{offset:1,color:'{sc}02'}}
          ])}}
        }}""")
        # PIB ref dashed gray
        series_js.append(f"""{{
          type:'line',xAxisIndex:{idx},yAxisIndex:{idx},data:pibTotal,
          symbol:'none',lineStyle:{{color:'#9CA3AF',width:1,type:'dashed'}},
          itemStyle:{{color:'#9CA3AF'}}
        }}""")
        # Zero line
        series_js.append(f"""{{
          type:'line',xAxisIndex:{idx},yAxisIndex:{idx},
          data:new Array(labels.length).fill(0),
          symbol:'none',lineStyle:{{color:dk()?'#475569':'#E2DDD3',width:0.5}},
          silent:true
        }}""")

        # Panel title
        tx = left + w / 2
        ty = top - 2.2
        titles_js.append(f"""{{
          text:'{sn}',left:'{tx:.1f}%',top:'{ty:.1f}%',textAlign:'center',
          textStyle:{{fontFamily:'Inter',fontSize:10,fontWeight:600,color:'{sc}'}}
        }}""")

    chart_js = f"""
    var labels={lbl};
    var pibTotal={pib_js};
    chart.setOption({{
      backgroundColor:'transparent',
      title:[{','.join(titles_js)}],
      tooltip:{{
        trigger:'axis',show:false
      }},
      grid:[{','.join(grids_js)}],
      xAxis:[{','.join(xaxes_js)}],
      yAxis:[{','.join(yaxes_js)}],
      series:[{','.join(series_js)}]
    }},true);"""

    html = populi_html(
        title="PIB Crecimiento por Sectores",
        chart_height=700,
        section_title="Crecimiento del PIB por actividad económica",
        section_subtitle="Variación interanual (%). Línea de color = sector, línea gris punteada = PIB total",
        accent_bar_color="#005F73",
        build_chart_js=chart_js,
    )
    (EMBED_DIR / "pib_crecimiento_sectores.html").write_text(html, encoding="utf-8")


def build_chart4():
    """Chart 4: Stacked bar contribution by sector + line for PIB total"""
    print("  [4] pib_contribucion_sectores.html")

    labels, data = parse_base2017_transposed(FILE_PIB_CONTRIBUCION, data_rows=list(range(13, 28)))
    _, pib_total, sectors = split_sector_data(data)

    lbl = json.dumps(labels)
    pib_js = json.dumps(json_safe(pib_total))

    bar_series = []
    for i, (sk, sv) in enumerate(sectors[:11]):
        dv = json.dumps(json_safe(sv))
        c = SECTOR_COLORS_LIST[i]
        n = SECTOR_SHORT[i]
        bar_series.append(f"""{{
          name:'{n}',type:'bar',stack:'total',data:{dv},
          itemStyle:{{color:'{c}'}},barMaxWidth:16
        }}""")

    # PIB total line on top
    bar_series.append(f"""{{
      name:'PIB Total',type:'line',data:{pib_js},
      symbol:'circle',symbolSize:4,z:10,
      lineStyle:{{color:dk()?'#E2E8F0':'#0D1B2A',width:2}},
      itemStyle:{{color:dk()?'#E2E8F0':'#0D1B2A'}}
    }}""")

    legend_names = json.dumps(SECTOR_SHORT[:11] + ["PIB Total"])

    chart_js = f"""
    var labels={lbl};
    chart.setOption({{
      backgroundColor:'transparent',
      grid:{{top:36,right:16,bottom:80,left:52}},
      tooltip:{js_tooltip_bar_stack('pp')},
      legend:{{
        data:{legend_names},
        bottom:4,type:'scroll',
        textStyle:{{fontFamily:'Inter',fontSize:10,color:dk()?'#64748B':'#64748B'}}
      }},
      xAxis:{js_xaxis_bar('labels',45,1)},
      yAxis:{js_yaxis('pp')},
      series:[{','.join(bar_series)}]
    }},true);"""

    html = populi_html(
        title="Contribución Sectorial al PIB",
        chart_height=480,
        section_title="Contribución al crecimiento del PIB",
        section_subtitle="Puntos porcentuales por actividad económica. Línea = crecimiento total del PIB",
        accent_bar_color="#EE9B00",
        build_chart_js=chart_js,
    )
    (EMBED_DIR / "pib_contribucion_sectores.html").write_text(html, encoding="utf-8")


def build_chart5():
    """Chart 5: 100% stacked area - sector shares"""
    print("  [5] pib_estructura_sectores.html")

    labels, data = parse_base2017_transposed(
        FILE_PIB_ESTRUCTURA, sheet_name="T_01.02", data_rows=list(range(13, 28)))
    _, _, sectors = split_sector_data(data)

    lbl = json.dumps(labels)

    area_series = []
    for i, (sk, sv) in enumerate(sectors[:11]):
        dv = json.dumps(json_safe(sv, 1))
        c = SECTOR_COLORS_LIST[i]
        n = SECTOR_SHORT[i]
        area_series.append(f"""{{
          name:'{n}',type:'line',stack:'total',data:{dv},
          symbol:'none',
          lineStyle:{{width:0.5,color:'{c}'}},
          itemStyle:{{color:'{c}'}},
          areaStyle:{{opacity:0.85}},
          emphasis:{{focus:'series'}}
        }}""")

    legend_names = json.dumps(SECTOR_SHORT[:11])

    chart_js = f"""
    var labels={lbl};
    chart.setOption({{
      backgroundColor:'transparent',
      grid:{{top:30,right:16,bottom:80,left:52}},
      tooltip:{{
        trigger:'axis',
        axisPointer:{{type:'cross'}},
        backgroundColor:dk()?'#080808':'rgba(13,27,42,.96)',
        borderColor:dk()?'#2A3A50':'#1A2940',
        borderWidth:1,
        textStyle:{{fontFamily:'Inter',fontSize:12,color:'#fff'}},
        padding:[12,16],
        extraCssText:'border-radius:10px;box-shadow:0 8px 32px rgba(0,0,0,.3);',
        formatter:function(params){{
          if(!params||!params.length) return '';
          var s='<div style="color:#E9D8A6;font-weight:600;margin-bottom:6px">'+params[0].axisValue+'</div>';
          params.forEach(function(p){{
            if(p.value!=null){{
              s+=p.marker+' <span style="color:#ccc">'+p.seriesName+'</span> '
                +'<span style="font-family:\\'JetBrains Mono\\';font-weight:700;color:#fff">'
                +p.value.toFixed(1)+' %</span><br/>';
            }}
          }});
          return s;
        }}
      }},
      legend:{{
        data:{legend_names},bottom:4,type:'scroll',
        textStyle:{{fontFamily:'Inter',fontSize:10,color:dk()?'#64748B':'#64748B'}}
      }},
      xAxis:{js_xaxis('labels',45,1)},
      yAxis:{{
        type:'value',max:100,
        splitLine:{{lineStyle:{{color:dk()?'#141414':'#F1F5F9',type:'dashed'}}}},
        axisLine:{{show:false}},axisTick:{{show:false}},
        axisLabel:{{fontFamily:'JetBrains Mono',fontSize:11,color:dk()?'#64748B':'#64748B',formatter:'{{value}}%'}}
      }},
      series:[{','.join(area_series)}]
    }},true);"""

    html = populi_html(
        title="Estructura del PIB por sector",
        chart_height=480,
        section_title="Estructura del PIB por actividad económica",
        section_subtitle="Participación porcentual a precios corrientes",
        accent_bar_color="#9B2226",
        build_chart_js=chart_js,
    )
    (EMBED_DIR / "pib_estructura_sectores.html").write_text(html, encoding="utf-8")


def build_chart6():
    """Chart 6: Stacked area - expenditure composition (chained Bs)"""
    print("  [6] pib_gasto_composicion.html")

    labels, data = parse_base2017_transposed(
        FILE_GASTO_NIVEL, data_rows=[13, 15, 16, 17, 18, 19, 20])

    keys = list(data.keys())
    series_parts = []  # (name, color, values, stack?)

    for k in keys:
        info = classify_gasto(k)
        if info is None:
            continue
        name, color = info
        if name == "PIB Total" or name == "FBCF":
            continue  # skip total and FBCF subset
        vals = json_safe(data[k], 1)
        if name == "Importaciones":
            vals = [(-v if v is not None else None) for v in vals]
            series_parts.append((f"Importaciones (−)", color, vals, False))
        else:
            series_parts.append((name, color, vals, True))

    lbl = json.dumps(labels)

    area_series = []
    for name, color, vals, is_stacked in series_parts:
        dv = json.dumps(vals)
        stack_str = "stack:'gasto'," if is_stacked else ""
        opacity = "0.7" if is_stacked else "0.15"
        dash = "" if is_stacked else "type:'dashed',"
        area_series.append(f"""{{
          name:'{name}',type:'line',data:{dv},{stack_str}
          symbol:'none',
          lineStyle:{{width:{'0.5' if is_stacked else '2'},color:'{color}',{dash}}},
          itemStyle:{{color:'{color}'}},
          areaStyle:{{opacity:{opacity}}}
        }}""")

    legend_names = json.dumps([s[0] for s in series_parts])

    chart_js = f"""
    var labels={lbl};
    chart.setOption({{
      backgroundColor:'transparent',
      grid:{{top:30,right:16,bottom:70,left:68}},
      tooltip:{{
        trigger:'axis',
        axisPointer:{{type:'cross'}},
        backgroundColor:dk()?'#080808':'rgba(13,27,42,.96)',
        borderColor:dk()?'#2A3A50':'#1A2940',
        borderWidth:1,
        textStyle:{{fontFamily:'Inter',fontSize:12,color:'#fff'}},
        padding:[12,16],
        extraCssText:'border-radius:10px;box-shadow:0 8px 32px rgba(0,0,0,.3);',
        formatter:function(params){{
          if(!params||!params.length) return '';
          var s='<div style="color:#E9D8A6;font-weight:600;margin-bottom:6px">'+params[0].axisValue+'</div>';
          params.forEach(function(p){{
            if(p.value!=null){{
              s+=p.marker+' <span style="color:#ccc">'+p.seriesName+'</span> '
                +'<span style="font-family:\\'JetBrains Mono\\';font-weight:700;color:#fff">'
                +p.value.toLocaleString('es-BO')+'</span> <span style="color:#999">MM</span><br/>';
            }}
          }});
          return s;
        }}
      }},
      legend:{{
        data:{legend_names},bottom:4,type:'scroll',
        textStyle:{{fontFamily:'Inter',fontSize:10,color:dk()?'#64748B':'#64748B'}}
      }},
      xAxis:{js_xaxis('labels',45,1)},
      yAxis:{{
        type:'value',name:'Millones Bs',
        nameTextStyle:{{fontFamily:'Inter',fontSize:10,color:dk()?'#64748B':'#64748B'}},
        splitLine:{{lineStyle:{{color:dk()?'#141414':'#F1F5F9',type:'dashed'}}}},
        axisLine:{{show:false}},axisTick:{{show:false}},
        axisLabel:{{fontFamily:'JetBrains Mono',fontSize:11,color:dk()?'#64748B':'#64748B',
          formatter:function(v){{return (v/1000).toFixed(0)+'k'}}
        }}
      }},
      series:[{','.join(area_series)}]
    }},true);"""

    html = populi_html(
        title="PIB por tipo de gasto",
        chart_height=460,
        section_title="PIB por tipo de gasto",
        section_subtitle="Millones de Bs encadenados (referencia 2017). Importaciones se restan",
        accent_bar_color="#0A9396",
        build_chart_js=chart_js,
    )
    (EMBED_DIR / "pib_gasto_composicion.html").write_text(html, encoding="utf-8")


def build_chart7():
    """Chart 7: Multi-line - growth rate by expenditure component"""
    print("  [7] pib_gasto_crecimiento.html")

    labels, data = parse_base2017_transposed(
        FILE_GASTO_CRECIMIENTO, data_rows=[13, 14, 15, 16, 17, 18, 19])

    keys = list(data.keys())

    lines = []
    for k in keys:
        info = classify_gasto(k)
        if info is None:
            continue
        name, color = info
        if name == "FBCF":
            continue
        vals = json_safe(data[k])
        width = 2.5 if name == "PIB Total" else 1.8
        dashed = False
        lines.append((name, color, vals, width, dashed))

    lbl = json.dumps(labels)

    line_series = []
    for name, color, vals, width, dashed in lines:
        dv = json.dumps(vals)
        dash_str = "type:'dashed'," if dashed else ""
        z = 10 if name == "PIB Total" else 2
        line_series.append(f"""{{
          name:'{name}',type:'line',data:{dv},
          symbol:'none',z:{z},
          lineStyle:{{width:{width},color:'{color}',{dash_str}}},
          itemStyle:{{color:'{color}'}},
          areaStyle:{{color:new echarts.graphic.LinearGradient(0,0,0,1,[
            {{offset:0,color:'{color}18'}},{{offset:1,color:'{color}02'}}
          ])}}
        }}""")

    # Zero reference
    line_series.append("""{{
      type:'line',data:new Array(labels.length).fill(0),
      symbol:'none',lineStyle:{{color:dk()?'#475569':'#E2DDD3',width:0.8}},
      silent:true,name:''
    }}""".replace("{{", "{").replace("}}", "}"))
    # Fix: the replace above won't work inside f-string, let's use a different approach

    legend_names = json.dumps([l[0] for l in lines])

    # Build zero line manually
    zero_line = """{
      type:'line',data:zeroLine,
      symbol:'none',lineStyle:{color:dk()?'#475569':'#E2DDD3',width:0.8},
      silent:true,name:''
    }"""

    chart_js = f"""
    var labels={lbl};
    var zeroLine=new Array(labels.length).fill(0);
    chart.setOption({{
      backgroundColor:'transparent',
      grid:{{top:30,right:16,bottom:70,left:52}},
      tooltip:{js_tooltip('%')},
      legend:{{
        data:{legend_names},bottom:4,type:'scroll',
        textStyle:{{fontFamily:'Inter',fontSize:10,color:dk()?'#64748B':'#64748B'}}
      }},
      xAxis:{js_xaxis('labels',45,1)},
      yAxis:{js_yaxis('%','{value}%')},
      series:[{','.join(line_series)},{zero_line}]
    }},true);"""

    html = populi_html(
        title="Crecimiento PIB por gasto",
        chart_height=460,
        section_title="Crecimiento del PIB por componente de gasto",
        section_subtitle="Variación interanual (%) de cada componente de gasto",
        accent_bar_color="#005F73",
        build_chart_js=chart_js,
    )
    (EMBED_DIR / "pib_gasto_crecimiento.html").write_text(html, encoding="utf-8")


def build_chart8():
    """Chart 8: Stacked bar contribution by expenditure"""
    print("  [8] pib_gasto_contribucion.html")

    labels, data = parse_base2017_transposed(
        FILE_GASTO_CONTRIBUCION, data_rows=[13, 15, 16, 17, 18, 19, 20])

    keys = list(data.keys())
    pib_total = None
    bar_parts = []

    for k in keys:
        info = classify_gasto(k)
        if info is None:
            continue
        name, color = info
        if name == "PIB Total":
            pib_total = json_safe(data[k])
            continue
        if name == "FBCF":
            continue
        vals = json_safe(data[k])
        if name == "Importaciones":
            vals = [(-v if v is not None else None) for v in vals]
            bar_parts.append((f"Importaciones (−)", color, vals))
        else:
            bar_parts.append((name, color, vals))

    lbl = json.dumps(labels)
    pib_js = json.dumps(pib_total) if pib_total else "[]"

    bar_series = []
    for name, color, vals in bar_parts:
        dv = json.dumps(vals)
        bar_series.append(f"""{{
          name:'{name}',type:'bar',stack:'total',data:{dv},
          itemStyle:{{color:'{color}'}},barMaxWidth:16
        }}""")

    # PIB total line overlay
    bar_series.append(f"""{{
      name:'PIB Total',type:'line',data:{pib_js},
      symbol:'circle',symbolSize:4,z:10,
      lineStyle:{{color:dk()?'#E2E8F0':'#0D1B2A',width:2}},
      itemStyle:{{color:dk()?'#E2E8F0':'#0D1B2A'}}
    }}""")

    legend_names = json.dumps([b[0] for b in bar_parts] + ["PIB Total"])

    chart_js = f"""
    var labels={lbl};
    chart.setOption({{
      backgroundColor:'transparent',
      grid:{{top:30,right:16,bottom:70,left:52}},
      tooltip:{js_tooltip_bar_stack('pp')},
      legend:{{
        data:{legend_names},bottom:4,type:'scroll',
        textStyle:{{fontFamily:'Inter',fontSize:10,color:dk()?'#64748B':'#64748B'}}
      }},
      xAxis:{js_xaxis_bar('labels',45,1)},
      yAxis:{js_yaxis('pp')},
      series:[{','.join(bar_series)}]
    }},true);"""

    html = populi_html(
        title="Contribución al PIB por gasto",
        chart_height=460,
        section_title="Contribución al crecimiento del PIB por tipo de gasto",
        section_subtitle="Puntos porcentuales. Línea = crecimiento total del PIB",
        accent_bar_color="#EE9B00",
        build_chart_js=chart_js,
    )
    (EMBED_DIR / "pib_gasto_contribucion.html").write_text(html, encoding="utf-8")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("POPULI - Generando charts PIB (Design System)")
    print(f"  Base: {BASE_DIR}")
    print(f"  Output: {EMBED_DIR}")
    print("=" * 60)

    build_chart1()
    build_chart2()
    build_chart3()
    build_chart4()
    build_chart5()
    build_chart6()
    build_chart7()
    build_chart8()

    print("=" * 60)
    print("Listo: 8 archivos HTML generados en embed/")
    print("=" * 60)


if __name__ == "__main__":
    main()
